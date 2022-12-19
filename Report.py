from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from Statistic import Statistics as st


class Report:
    def __init__(self, dataYear, dataCity, profession, vacancies):
        self.DataYear = dataYear
        self.DataCity = dataCity
        self.Profession = profession
        self.vacancies = vacancies

    def generateExcel(self):
        wb = Workbook()
        self.SetSheetYear(wb.create_sheet('статистика по годам', 0))
        self.SetSheetCity(wb.create_sheet('статиситка по городам', 1))
        wb.save("данные вакансий")

    def SetSheetYear(self, wsDataYear):
        wsDataYear['A1'] = 'Год'
        wsDataYear['B1'] = 'средняя зарплата'
        wsDataYear['C1'] = f'средняя зарплата - {self.Profession}'
        wsDataYear['D1'] = 'количество вакансий'
        wsDataYear['E1'] = f'количество вакансий -{self.Profession}'
        start = 2
        for k, v in st.GetDynamicsSalaryByYear(self.DataYear).items():
            wsDataYear[f'A{start}'] = k
            wsDataYear[f'B{start}'] = v
            # wsDataYear[f'A{i}'] =
            # wsDataYear[f'D{i}'] = len(v)
            # wsDataYear[f'A{i}'] =
            start += 1
        wsDataYear.column_dimensions['A'].width = 5
        wsDataYear.column_dimensions['B'].width = 18
        wsDataYear.column_dimensions['C'].width = 28
        wsDataYear.column_dimensions['E'].width = 28
        wsDataYear.column_dimensions['D'].width = 22

        start = 2
        for k, v in st.GetDynamicsWithFilterSalaryByYear(self.DataYear, self.Profession).items():
            wsDataYear[f'C{start}'] = v
            start += 1
        start = 2
        for k, v in st.GetDynamicsVacancyByYear(self.DataYear).items():
            wsDataYear[f'E{start}'] = v
            start += 1
        start = 2
        for k, v in st.GetDynamicsWithFilterVacancyByYear(self.DataYear, self.Profession).items():
            wsDataYear[f'D{start}'] = v
            start += 1

    def SetSheetCity(self, wsDataCity):
        wsDataCity['A1'] = 'Город'
        wsDataCity['B1'] = 'Уровень зарплат'
        wsDataCity['D1'] = 'Город'
        wsDataCity['E1'] = 'Доля вакансий'
        wsDataCity.column_dimensions['A'].width = 7
        wsDataCity.column_dimensions['B'].width = 17
        wsDataCity.column_dimensions['D'].width = 7
        wsDataCity.column_dimensions['E'].width = 15
        start = 2
        for k, v in st.GetDynamicsSalaryByCity(self.DataYear, self.vacancies).items():
            wsDataCity[f'A{start}'] = k
            wsDataCity[f'B{start}'] = v
            start += 1
        start = 2

        for k, v in st.GetDynamicsProcentCountByCity(self.DataYear, self.vacancies).items():
            wsDataCity[f'D{start}'] = k
            wsDataCity[f'E{start}'] = v
            start += 1